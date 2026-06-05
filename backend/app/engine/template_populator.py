from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from .sf_client import SuccessFactorsClient


FOUNDATION_DATA_SHEETS = {
    "Legal Entity Data": {
        "entity": "FOCompany",
        "fields": {
            "legal entity code": "externalCode",
            "legal entity name": "name",
            "legal entity description": "description",
            "country": "country",
            "currency": "currency",
            "standard weekly hours": "standardHours",
            "default pay group": "defaultPayGroup",
            "default location": "defaultLocation",
            "name format": "nameFormat",
        },
    },
    "Business Unit Data": {
        "entity": "FOBusinessUnit",
        "fields": {
            "business unit code": "externalCode",
            "business unit name": "name",
            "business unit description": "description",
            "head of unit": "headOfUnit",
        },
    },
    "Division Data": {
        "entity": "FODivision",
        "fields": {
            "division code": "externalCode",
            "division name": "name",
            "division description": "description",
            "head of division": "headOfUnit",
            "parent business unit": "parentBusinessUnit",
        },
    },
    "Department Data": {
        "entity": "FODepartment",
        "fields": {
            "department code": "externalCode",
            "department name": "name",
            "department description": "description",
            "head of department": "headOfUnit",
            "parent department": "parent",
            "cost centre": "costCenter",
            "parent division": "parentDivision",
        },
    },
    "Location Data": {
        "entity": "FOLocation",
        "fields": {
            "location code": "externalCode",
            "location name": "name",
            "location description": "description",
            "location group": "locationGroup",
            "standard hours": "standardHours",
            "timezone": "timezone",
            "geozone": "geozone",
            "parent legal entity": "company",
            "address line 1": "address1",
            "address line 2": "address2",
            "city": "city",
            "country": "country",
        },
    },
    "Location Group Data": {
        "entity": "FOLocationGroup",
        "fields": {
            "location group code": "externalCode",
            "location group name": "name",
            "location group description": "description",
        },
    },
    "Cost Centre Data": {
        "entity": "FOCostCenter",
        "fields": {
            "cost centre code": "externalCode",
            "cost centre name": "name",
            "cost centre description": "description",
            "parent cost centre": "parent",
            "cost centre manager": "costCenterManager",
            "gl statement code": "glStatementCode",
            "cost centre external object id": "costCenterExternalObjectId",
            "parent legal entity": "company",
        },
    },
    "Job Function Data": {
        "entity": "FOJobFunction",
        "fields": {
            "job function code": "externalCode",
            "job function description": "name",
            "parent job family": "parentJobFamily",
            "job family type": "jobFamilyType",
        },
    },
    "Job Classification Data": {
        "entity": "FOJobCode",
        "fields": {
            "job code code": "externalCode",
            "job code name": "name",
            "job code description": "description",
            "pay grade": "payGrade",
            "job family": "jobFunction",
            "job level": "jobLevel",
            "parent job classification": "parentJobCode",
        },
    },
    "Pay Group Data": {
        "entity": "FOPayGroup",
        "fields": {
            "pay group code": "externalCode",
            "pay group name": "name",
            "pay group description": "description",
            "parent legal entity": "company",
        },
    },
    "Pay Grade Data": {
        "entity": "FOPayGrade",
        "fields": {
            "pay grade code": "externalCode",
            "pay grade name": "name",
            "pay grade description": "description",
        },
    },
    "Frequency Data": {
        "entity": "FOFrequency",
        "fields": {
            "frequency code": "externalCode",
            "frequency name": "name",
            "frequency description": "description",
            "annualisation factor": "annualizationFactor",
        },
    },
}


def populate_foundation_template(
    *,
    template_path: Path,
    output_dir: Path,
    client: SuccessFactorsClient,
    top: int = 200,
) -> tuple[Path, list[dict[str, Any]]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "Template population needs openpyxl. Install requirements or run with the bundled Python runtime."
        ) from exc

    output_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(template_path)
    status_rows: list[dict[str, Any]] = []

    for sheet_name, spec in FOUNDATION_DATA_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            status_rows.append(
                {"sheet": sheet_name, "entity": spec["entity"], "status": "Missing sheet", "rows": 0}
            )
            continue

        ws = wb[sheet_name]
        try:
            records = client.fetch_entity_records(spec["entity"], top=top)
            written = _populate_data_sheet(ws, records, spec["fields"])
            status_rows.append(
                {"sheet": sheet_name, "entity": spec["entity"], "status": "Populated", "rows": written}
            )
        except Exception as exc:
            status_rows.append(
                {
                    "sheet": sheet_name,
                    "entity": spec["entity"],
                    "status": "Error",
                    "rows": 0,
                    "error": str(exc),
                }
            )

    _write_run_summary(wb, status_rows)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"populated_ec_foundation_{timestamp}.xlsx"
    wb.save(output_path)
    return output_path, status_rows


def _populate_data_sheet(ws, records: list[dict], field_map: dict[str, str]) -> int:
    header_row = 3
    start_row = 4
    headers = [
        _normalize(ws.cell(row=header_row, column=col).value)
        for col in range(1, ws.max_column + 1)
    ]
    _clear_existing_values(ws, start_row)
    _ensure_rows(ws, start_row, len(records))

    for row_offset, record in enumerate(records):
        row_index = start_row + row_offset
        for col_index, header in enumerate(headers, start=1):
            field_path = field_map.get(header)
            if field_path:
                ws.cell(row=row_index, column=col_index, value=_stringify(_dig(record, field_path)))
    return len(records)


def _clear_existing_values(ws, start_row: int) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.value = None


def _ensure_rows(ws, start_row: int, needed: int) -> None:
    existing = max(ws.max_row - start_row + 1, 0)
    if needed <= existing:
        return
    template_row = start_row if ws.max_row >= start_row else start_row - 1
    for _ in range(needed - existing):
        ws.append([])
        source_row = template_row
        target_row = ws.max_row
        for col in range(1, ws.max_column + 1):
            source = ws.cell(row=source_row, column=col)
            target = ws.cell(row=target_row, column=col)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            if source.alignment:
                target.alignment = copy(source.alignment)


def _write_run_summary(wb, status_rows: list[dict[str, Any]]) -> None:
    title = "POC Run Summary"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title, 0)
    ws.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([])
    ws.append(["Sheet", "OData Entity", "Status", "Rows Written", "Error"])
    for item in status_rows:
        ws.append(
            [
                item.get("sheet", ""),
                item.get("entity", ""),
                item.get("status", ""),
                item.get("rows", 0),
                item.get("error", ""),
            ]
        )
    ws.freeze_panes = "A4"
    for column in ws.columns:
        ws.column_dimensions[column[0].column_letter].width = min(
            max(len(str(cell.value or "")) for cell in column) + 2,
            80,
        )


def _dig(record: dict, field_path: str) -> Any:
    value: Any = record
    for part in field_path.split("."):
        if not isinstance(value, dict):
            return ""
        value = value.get(part)
    if isinstance(value, dict):
        for key in ("externalCode", "code", "name", "defaultValue", "label"):
            if key in value:
                return value[key]
        return ""
    return value


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str) and value.startswith("/Date("):
        return value
    return str(value)


def _normalize(value: Any) -> str:
    text = str(value or "").replace("\xa0", " ").strip().lower()
    return " ".join(text.split())
