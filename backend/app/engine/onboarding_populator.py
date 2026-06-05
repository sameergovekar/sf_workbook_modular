from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl.cell.cell import MergedCell

from .sf_client import SuccessFactorsClient, list_entity_type_names


ONBOARDING_SHEETS = {
    "Responsible Groups": {
        "entity": "DynamicGroupDefinition",
        "entities": [
            "DynamicGroupDefinition",
            "ONB2BuddyActivityResponsible",
            "ONB2EquipmentActivityResponsible",
        ],
        "search_terms": ["responsible", "group", "onb"],
        "header_row": 4,
        "start_row": 5,
        "fields": {
            "assignment order": "assignmentOrder",
            "responsible group": "name",
            "assignment type": "groupType",
            "assignment type (user, role, dynamic group)": "assignmentType",
            "dynamic group name": "dynamicGroup",
            "group members": "groupMembers",
            "individual users": "users",
            "target group in business rule": "targetGroup",
            "comments": "description",
        },
    },
    "Business Rules": {
        "entity": "RBPRule",
        "entities": [
            "RBPRule",
            "EventDeterminationRuleConfiguration",
            "LegalEntityListForEventRule",
        ],
        "search_terms": ["rule", "business"],
        "header_row": 4,
        "start_row": 5,
        "fields": {
            "name": "ruleName",
            "purpose": "description",
            "what should the rule state?": "ruleType",
            "if statement": "baseObject",
            "then statement": "scenario",
        },
    },
    "ONB Picklists": {
        "entity": "PickListValueV2",
        "entities": [
            "PickListValueV2",
            "PickListV2",
            "Picklist",
        ],
        "search_terms": ["picklist", "picklistvalue"],
        "header_row": 5,
        "start_row": 6,
        "fields": {
            "* external code": "externalCode",
            "* label": "label.defaultValue",
        },
    },
}


ONBOARDING_COVERAGE = {
    "Cover": {
        "support": "Manual",
        "reason": "Cover page and project/client branding are not API-backed configuration.",
        "next_action": "Fill manually or derive from project metadata later.",
    },
    " Configuration Survey": {
        "support": "Manual",
        "reason": "This is a business/design questionnaire, not a direct SuccessFactors OData object.",
        "next_action": "Keep as workshop input.",
    },
    "Introduction": {
        "support": "Manual",
        "reason": "Static implementation guidance text.",
        "next_action": "Leave unchanged.",
    },
    "Instance Info": {
        "support": "Auto",
        "reason": "Can be populated from API URL/company ID/local runtime inputs.",
        "next_action": "Already populated by the POC.",
    },
    "Compliance": {
        "support": "Partial",
        "reason": "Some compliance form availability may be API-backed, but country/legal-entity decisions are business inputs.",
        "next_action": "Map available compliance entities after tenant discovery.",
    },
    "NY WTPA": {
        "support": "Manual",
        "reason": "Jurisdiction-specific business decision sheet.",
        "next_action": "Keep as manual input unless a specific compliance API is identified.",
    },
    "Processes": {
        "support": "Partial",
        "reason": "Process templates may be discoverable, but final process applicability is a design decision.",
        "next_action": "Search for process/task entities and map if exposed.",
    },
    "Hire Template": {
        "support": "Partial",
        "reason": "Some hire template/block metadata may be API-backed, but labels/visibility often need configuration-specific mapping.",
        "next_action": "Search metadata for hire template and data collection entities.",
    },
    "Onboarding Programs": {
        "support": "Partial",
        "reason": "Program/task objects may be exposed, but program conditions and rule logic need mapping.",
        "next_action": "Search for onboarding process/program/task entities.",
    },
    "Custom Tasks": {
        "support": "Partial",
        "reason": "Custom MDF/task objects may be exposed if configured in the tenant.",
        "next_action": "Map task entities after discovery.",
    },
    "Email Services": {
        "support": "Partial",
        "reason": "Notification templates/triggers may be exposed, but custom layout/activation decisions need validation.",
        "next_action": "Search for notification/email template entities.",
    },
    "Home Page Cards ": {
        "support": "Partial",
        "reason": "Home page card configuration may be available in platform APIs, but template columns are design-oriented.",
        "next_action": "Discover card/homepage entities.",
    },
    "Forms & Policies": {
        "support": "Partial",
        "reason": "Document/form metadata may be discoverable, but source PDFs and signing decisions are project inputs.",
        "next_action": "Search for document/form/policy entities.",
    },
    "Compliance Forms": {
        "support": "Partial",
        "reason": "Supported forms are partly standard content; activation per country/legal entity needs tenant-specific mapping.",
        "next_action": "Map compliance form entities if exposed.",
    },
    "Custom ONB Data Collection": {
        "support": "Partial",
        "reason": "Custom data collection MDF objects may be API-backed, but form design needs mapping.",
        "next_action": "Search for custom data collection object definitions.",
    },
    "ONB Picklists": {
        "support": "Auto",
        "reason": "Tenant exposes picklist candidates such as PickListValueV2.",
        "next_action": "Already attempted by the POC.",
    },
    "Business Rules": {
        "support": "Auto",
        "reason": "Tenant exposes rule candidates such as RBPRule.",
        "next_action": "Already attempted by the POC; field mapping may need refinement.",
    },
    "Data Protection & Privacy": {
        "support": "Manual",
        "reason": "Mostly policy/legal decisions and retention requirements.",
        "next_action": "Keep as manual input unless DPP configuration API is identified.",
    },
    "Responsible Groups": {
        "support": "Auto",
        "reason": "Tenant exposes group/responsible candidates such as DynamicGroupDefinition.",
        "next_action": "Already attempted by the POC; field mapping may need refinement.",
    },
    "Permission Roles": {
        "support": "Partial",
        "reason": "RBP roles are likely API-backed, but target populations and role design need mapping.",
        "next_action": "Search for permission role entities.",
    },
    "Permission Groups": {
        "support": "Partial",
        "reason": "RBP groups are likely API-backed, but proxy/service-user decisions are manual.",
        "next_action": "Search for permission group entities.",
    },
    "Role Permission Settings": {
        "support": "Partial",
        "reason": "Permissions may be exported, but the matrix is large and role-specific.",
        "next_action": "Map after identifying RBP permission entities.",
    },
    "Recruit-to-Hire Mapping ": {
        "support": "Partial",
        "reason": "Some integration/template mapping may be API-backed; business mapping decisions still need review.",
        "next_action": "Search for recruit-to-hire/integration mapping entities.",
    },
    "Reporting Requirements": {
        "support": "Manual",
        "reason": "Reporting requirements are project/business requirements, not tenant configuration extracts.",
        "next_action": "Keep as manual workshop input.",
    },
    "US State Tax Codes": {
        "support": "Manual",
        "reason": "Static reference sheet for US tax mapping.",
        "next_action": "Leave unchanged unless SAP exposes updated tax reference data.",
    },
    "US State W4 Tax Info": {
        "support": "Manual",
        "reason": "Static reference sheet for state W-4 fields.",
        "next_action": "Leave unchanged unless SAP exposes updated tax reference data.",
    },
    "US Federal W4 Info": {
        "support": "Manual",
        "reason": "Static reference sheet for federal W-4 mapping.",
        "next_action": "Leave unchanged unless SAP exposes updated tax reference data.",
    },
    "E-Verify Worksheet": {
        "support": "Manual",
        "reason": "Customer/company setup worksheet, not a direct OData extract.",
        "next_action": "Keep as manual input.",
    },
    "Cutover Tasks": {
        "support": "Manual",
        "reason": "Project plan/checklist rather than SuccessFactors configuration.",
        "next_action": "Keep as manual project planning input.",
    },
    "Instance Sync": {
        "support": "Partial",
        "reason": "Some syncable objects can be discovered, but overwrite decisions and cutover tracking are manual.",
        "next_action": "Use discovery output to propose sync object list.",
    },
    "Source-Do not change": {
        "support": "Manual",
        "reason": "Static source/reference sheet.",
        "next_action": "Leave unchanged.",
    },
}


def populate_onboarding_template(
    *,
    template_path: Path,
    output_dir: Path,
    client: SuccessFactorsClient,
    top: int = 200,
) -> tuple[Path, list[dict[str, Any]]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "Onboarding template population needs openpyxl. Run with the bundled Python runtime."
        ) from exc

    output_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(template_path)
    status_rows: list[dict[str, Any]] = []
    entity_names: list[str] | None = None

    _populate_instance_info(wb, client.config)
    status_rows.append(
        {"sheet": "Instance Info", "entity": "Local input", "status": "Populated", "rows": 1}
    )

    for sheet_name, spec in ONBOARDING_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            status_rows.append(
                {"sheet": sheet_name, "entity": spec["entity"], "status": "Missing sheet", "rows": 0}
            )
            continue

        ws = wb[sheet_name]
        try:
            entity, records = _fetch_first_available(client, spec, top)
            if sheet_name == "ONB Picklists":
                records = _filter_onboarding_picklists(records)
            written = _populate_mapped_sheet(
                ws=ws,
                records=records,
                header_row=spec["header_row"],
                start_row=spec["start_row"],
                field_map=spec["fields"],
            )
            status = "Populated"
            error = ""
            if records and written == 0:
                status = "Fetched - no mapped workbook fields"
                error = f"Entity returned {len(records)} records, but mapped fields did not match workbook columns."
            status_rows.append(
                {
                    "sheet": sheet_name,
                    "entity": entity,
                    "status": status,
                    "rows": written,
                    "error": error,
                }
            )
        except Exception as exc:
            entity_names = entity_names or _safe_entity_names(client)
            candidates = _candidate_entities(entity_names, spec.get("search_terms", []))
            status_rows.append(
                {
                    "sheet": sheet_name,
                    "entity": spec["entity"],
                    "status": "Error",
                    "rows": 0,
                    "error": f"{exc}. Candidate entities: {', '.join(candidates[:12]) or 'none found'}",
                }
            )

    _write_coverage_sheet(wb)
    _write_run_summary(wb, status_rows)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"populated_onboarding_{timestamp}.xlsx"
    wb.save(output_path)
    return output_path, status_rows


def _write_coverage_sheet(wb) -> None:
    title = "POC Coverage Map"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title, 1)
    ws.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Purpose", "Shows which Onboarding workbook tabs are currently auto-populated, partially supportable, manual, or unknown."])
    ws.append([])
    ws.append(["Workbook Tab", "Support Level", "Reason", "Next Action"])

    for sheet_name in wb.sheetnames:
        if sheet_name in {"POC Run Summary", "POC Coverage Map"}:
            continue
        coverage = ONBOARDING_COVERAGE.get(
            sheet_name,
            {
                "support": "Unknown",
                "reason": "This tab has not been assessed yet.",
                "next_action": "Review workbook purpose and search tenant metadata for a matching entity.",
            },
        )
        ws.append(
            [
                sheet_name,
                coverage["support"],
                coverage["reason"],
                coverage["next_action"],
            ]
        )

    ws.freeze_panes = "A5"
    widths = {"A": 34, "B": 18, "C": 80, "D": 70}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = copy(cell.alignment)
            cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")


def _fetch_first_available(
    client: SuccessFactorsClient, spec: dict[str, Any], top: int
) -> tuple[str, list[dict]]:
    errors = []
    for entity in spec.get("entities", [spec["entity"]]):
        try:
            return entity, client.fetch_entity_records(entity, top=top)
        except Exception as exc:
            errors.append(f"{entity}: {exc}")
    raise RuntimeError("; ".join(errors))


def _safe_entity_names(client: SuccessFactorsClient) -> list[str]:
    try:
        return list_entity_type_names(client.fetch_service_metadata())
    except Exception:
        return []


def _candidate_entities(entity_names: list[str], terms: list[str]) -> list[str]:
    lowered_terms = [term.lower() for term in terms]
    scored = []
    for name in entity_names:
        lower_name = name.lower()
        score = sum(1 for term in lowered_terms if term in lower_name)
        if score:
            scored.append((score, name))
    return [name for _score, name in sorted(scored, key=lambda item: (-item[0], item[1].lower()))]


def _populate_instance_info(wb, config) -> None:
    if "Instance Info" not in wb.sheetnames:
        return
    ws = wb["Instance Info"]
    host = config.api_base_url.replace("https://", "").replace("http://", "").split("/", 1)[0]
    values = ["api44preview", "Preview", "Preview", "POC", config.company_id or host]
    for col_index, value in enumerate(values, start=1):
        ws.cell(row=4, column=col_index, value=value)


def _populate_mapped_sheet(
    *,
    ws,
    records: list[dict],
    header_row: int,
    start_row: int,
    field_map: dict[str, str],
) -> int:
    headers = [
        _normalize(ws.cell(row=header_row, column=col).value)
        for col in range(1, ws.max_column + 1)
    ]
    _clear_existing_values(ws, start_row)
    _ensure_rows(ws, start_row, len(records))

    rows_with_values = 0
    for row_offset, record in enumerate(records):
        row_index = start_row + row_offset
        row_has_value = False
        for col_index, header in enumerate(headers, start=1):
            field_path = field_map.get(header)
            if field_path:
                cell = ws.cell(row=row_index, column=col_index)
                if not isinstance(cell, MergedCell):
                    value = _stringify(_dig(record, field_path))
                    cell.value = value
                    row_has_value = row_has_value or bool(value)
        if row_has_value:
            rows_with_values += 1
    return rows_with_values


def _filter_onboarding_picklists(records: list[dict]) -> list[dict]:
    filtered = []
    for record in records:
        picklist_id = _stringify(_dig(record, "picklist.id") or _dig(record, "id"))
        external_code = _stringify(record.get("externalCode", ""))
        if "ONB" in picklist_id.upper() or "ONB" in external_code.upper():
            filtered.append(record)
    return filtered or records


def _clear_existing_values(ws, start_row: int) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None


def _ensure_rows(ws, start_row: int, needed: int) -> None:
    existing = max(ws.max_row - start_row + 1, 0)
    if needed <= existing:
        return
    template_row = start_row if ws.max_row >= start_row else start_row - 1
    for _ in range(needed - existing):
        ws.append([])
        target_row = ws.max_row
        for col in range(1, ws.max_column + 1):
            source = ws.cell(row=template_row, column=col)
            target = ws.cell(row=target_row, column=col)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            if source.alignment:
                target.alignment = copy(source.alignment)


def _write_run_summary(wb, status_rows: list[dict[str, Any]]) -> None:
    title = "POC Run Summary"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title, 0)
    ws.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Template Mode", "Onboarding"])
    ws.append([])
    ws.append(["Sheet", "OData Entity", "Status", "Rows Written", "Error"])
    for item in status_rows:
        ws.append(
            [
                item.get("sheet", ""),
                item.get("entity", ""),
                item.get("status", ""),
                item.get("rows", 0),
                item.get("error", ""),
            ]
        )
    ws.freeze_panes = "A5"
    for column in ws.columns:
        ws.column_dimensions[column[0].column_letter].width = min(
            max(len(str(cell.value or "")) for cell in column) + 2,
            80,
        )


def _dig(record: dict, field_path: str) -> Any:
    value: Any = record
    for part in field_path.split("."):
        if not isinstance(value, dict):
            return ""
        value = value.get(part)
    if isinstance(value, dict):
        for key in ("defaultValue", "externalCode", "code", "name", "label"):
            if key in value:
                return value[key]
        return ""
    if isinstance(value, list):
        return ", ".join(_stringify(item) for item in value)
    return value


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dict):
        return _stringify(_dig(value, "defaultValue"))
    return str(value)


def _normalize(value: Any) -> str:
    text = str(value or "").replace("\xa0", " ").replace("\n", " ").strip().lower()
    text = text.replace("  ", " ")
    return " ".join(text.split())
